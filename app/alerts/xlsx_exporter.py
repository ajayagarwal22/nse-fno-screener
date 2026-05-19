"""Generate a comprehensive signals + paper trades Excel workbook.

Sheets:
  1. Signal × Trade  — every signal with all indicator values and gate results,
                        joined to its paper trade outcome (if any).
  2. Gate Reference  — static column-by-column explanation of all 9 gates and
                        the indicators/timeframes/thresholds behind each one.
  3. Paper Trades    — full chronological paper trade ledger.
  4. P&L Summary     — daily and overall performance statistics.
"""
import csv
import json
import sqlite3
from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

_DB_PATH = Path("/Users/ajayagarwal/nse-fno-screener-data/screener_trades.db")
_EXPORTS_DIR = Path(__file__).parent.parent.parent / "exports"

# ── Gate metadata ────────────────────────────────────────────────────────────

_GATES = [
    # (db_key,                     label,              weight, direction,   indicator,         timeframe,  threshold)
    ("regime_supportive",          "Regime Support",   10,     "CALL/PUT",  "Market Regime (VIX + A/D ratio)",  "Daily",   "Not bearish/event-risk/high-vol"),
    ("rs_positive",                "RS Positive",       5,     "CALL",      "Relative Strength vs Nifty",        "Daily",   "Stock return > Nifty by >0.01"),
    ("rs_negative",                "RS Negative",       5,     "PUT",       "Relative Strength vs Nifty",        "Daily",   "Stock return < Nifty by >0.01"),
    ("rsi_divergence",             "RSI Divergence",   30,     "CALL/PUT",  "RSI 14",                            "15-min",  "Bullish/bearish RSI divergence detected"),
    ("htf_trend_bullish",          "HTF Trend Bull",   15,     "CALL",      "EMA 21 / 50 / 200",                 "Daily",   "All EMAs aligned upward"),
    ("htf_trend_bearish",          "HTF Trend Bear",   15,     "PUT",       "EMA 21 / 50 / 200",                 "Daily",   "All EMAs aligned downward"),
    ("macd_bull_cross",            "MACD Bull Cross",  15,     "CALL",      "MACD (12, 26, 9)",                  "5-min",   "MACD line crosses above signal line"),
    ("macd_bear_cross",            "MACD Bear Cross",  15,     "PUT",       "MACD (12, 26, 9)",                  "5-min",   "MACD line crosses below signal line"),
    ("above_vwap",                 "Above VWAP",       10,     "CALL",      "VWAP",                              "5-min",   "Close > VWAP (intraday)"),
    ("below_vwap",                 "Below VWAP",       10,     "PUT",       "VWAP",                              "5-min",   "Close < VWAP (intraday)"),
    ("volume_expansion",           "Volume Expansion",  5,     "CALL/PUT",  "Volume vs 20-day avg",              "5-min",   "Current volume ≥ 1.5× 20-day avg"),
    ("put_writing_or_short_covering", "OI: Put Writing", 5,    "CALL",      "OI + OI Change (option chain)",     "Live",    "PE OI decreasing → short covering / put writers exiting"),
    ("call_writing_or_short_buildup", "OI: Call Writing", 5,   "PUT",       "OI + OI Change (option chain)",     "Live",    "CE OI increasing → call writers building short"),
    ("pcr_supportive",             "PCR Bullish",       5,     "CALL",      "Put/Call Ratio",                    "Live",    "PCR > 1.2 → more puts than calls"),
    ("pcr_bearish",                "PCR Bearish",       5,     "PUT",       "Put/Call Ratio",                    "Live",    "PCR < 0.8 → more calls than puts"),
]

# Subset used for the gate columns in Sheet 1 (direction-neutral names)
_SHEET1_GATES = [
    "regime_supportive",
    "rs_positive", "rs_negative",
    "rsi_divergence",
    "htf_trend_bullish", "htf_trend_bearish",
    "macd_bull_cross", "macd_bear_cross",
    "above_vwap", "below_vwap",
    "volume_expansion",
    "put_writing_or_short_covering", "call_writing_or_short_buildup",
    "pcr_supportive", "pcr_bearish",
]

# ── Colours ──────────────────────────────────────────────────────────────────

_BG_HEADER    = "1A1F2E"
_BG_SUBHEADER = "0F1520"
_BG_WIN       = "0A2E1A"
_BG_LOSS      = "2E0A0A"
_BG_EVEN      = "111827"
_BG_ODD       = "0D1117"
_BG_GATE_YES  = "0A2416"
_BG_GATE_NO   = "2A1010"
_BG_GATE_NA   = "1A1A2A"

_FG_HEADER    = "E2E8F0"
_FG_GREEN     = "22C55E"
_FG_RED       = "EF4444"
_FG_AMBER     = "F59E0B"
_FG_PURPLE    = "A78BFA"
_FG_MUTED     = "64748B"
_FG_WHITE     = "F1F5F9"

_THIN_BORDER_COLOUR = "1E2533"


def _font(bold=False, size=10, colour=_FG_WHITE):
    return Font(bold=bold, size=size, color=colour, name="Calibri")

def _fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def _border():
    s = Side(style="thin", color=_THIN_BORDER_COLOUR)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _write_header_row(ws, col_defs, row=1):
    """Write one header row from a list of (label, width) tuples."""
    for col_idx, (label, width) in enumerate(col_defs, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font      = _font(bold=True, size=10, colour=_FG_HEADER)
        cell.fill      = _fill(_BG_HEADER)
        cell.alignment = _align("center", wrap=True)
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 36


# ── Data loaders ─────────────────────────────────────────────────────────────

def _load_db_signals_and_trades():
    """Return (signals_rows, trades_by_signal_id) from the paper_trader SQLite DB."""
    if not _DB_PATH.exists():
        return [], {}

    conn = sqlite3.connect(str(_DB_PATH))
    conn.row_factory = sqlite3.Row

    signals = conn.execute(
        "SELECT * FROM signals ORDER BY timestamp"
    ).fetchall()

    trades = conn.execute(
        "SELECT * FROM trades ORDER BY entry_time"
    ).fetchall()

    trades_by_sid = {}
    for t in trades:
        sid = t["signal_id"]
        trades_by_sid.setdefault(sid, []).append(dict(t))

    conn.close()
    return [dict(s) for s in signals], trades_by_sid


def _load_csv_signals():
    """Load all signal CSVs from exports/. Returns a list of dicts, newest last."""
    rows = []
    for path in sorted(_EXPORTS_DIR.glob("signals_*.csv")):
        with open(path, newline="") as f:
            reader = csv.DictReader(f)
            for r in reader:
                r["_source"] = path.stem
                rows.append(r)
    return rows


# ── Sheet 1: Signal × Trade ───────────────────────────────────────────────────

_S1_COLS = [
    # (header label,                    width)
    ("Date",                             11),
    ("Time (IST)",                        9),
    ("Symbol",                           13),
    ("Direction",                        10),
    ("Confidence",                       11),
    ("Gate Score",                       11),
    # ── Layer 1 ──
    ("Regime Support\n(10 pts)",         14),
    # ── Layer 2 ──
    ("RS +/−\n(5 pts)",                  12),
    # ── Layer 3 ──
    ("RSI Divergence\n(30 pts)",         15),
    ("HTF Trend\n(15 pts)",              13),
    ("MACD Cross\n(15 pts)",             13),
    ("VWAP Side\n(10 pts)",              12),
    ("Volume Exp.\n(5 pts)",             12),
    # ── Layer 4 ──
    ("OI Writing\n(5 pts)",             12),
    ("PCR Gate\n(5 pts)",               11),
    # ── Indicator values ──
    ("Regime",                           16),
    ("VIX",                               7),
    ("RSI",                               7),
    ("MACD Hist",                        11),
    ("VWAP",                             10),
    ("PCR",                               7),
    ("OI Interp.",                       14),
    ("HTF Trend",                        11),
    ("Divergence",                       11),
    ("Candle Pattern\n(15-min)",         18),
    # ── Entry/exit levels ──
    ("Entry Spot",                       11),
    ("Stop Loss\n(spot)",                12),
    ("Target 1\n(spot)",                 12),
    ("Target 2\n(spot)",                 12),
    ("R:R",                               8),
    ("Position Size",                    15),
    # ── Paper trade ──
    ("Strike",                            9),
    ("Expiry",                           11),
    ("Opt Type",                          9),
    ("Lots",                              8),
    ("Entry Prem.",                      12),
    ("Invested Amt.\n(lots × prem)",     14),
    ("Entry Spot\n(trade)",              13),
    ("Entry Time",                       17),
    ("Exit Prem.",                       11),
    ("Exit Spot",                        11),
    ("Exit Time",                        17),
    ("Exit Reason",                      12),
    ("P&L Pts",                          10),
    ("P&L ₹",                            11),
    ("P&L %",                             9),
    ("Outcome",                          11),
]


def _gate_cell(ws, row, col, passed, label=""):
    if passed is None:
        val   = "—"
        bg    = _BG_GATE_NA
        fg    = _FG_MUTED
    elif passed:
        val   = f"✓  {label}" if label else "✓"
        bg    = _BG_GATE_YES
        fg    = _FG_GREEN
    else:
        val   = f"✗  {label}" if label else "✗"
        bg    = _BG_GATE_NO
        fg    = _FG_RED
    c = ws.cell(row=row, column=col, value=val)
    c.font      = _font(bold=passed is True, colour=fg)
    c.fill      = _fill(bg)
    c.alignment = _align("center")
    c.border    = _border()


def _build_sheet1(wb, signals, trades_by_sid):
    ws = wb.create_sheet("Signal × Trade Detail")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    _write_header_row(ws, _S1_COLS, row=1)

    for r_idx, sig in enumerate(signals, start=2):
        bg = _BG_EVEN if r_idx % 2 == 0 else _BG_ODD

        active_gates = set(json.loads(sig.get("active_signals") or "[]"))
        direction    = sig.get("direction", "")

        # Determine which rs gate to check
        rs_gate = "rs_positive" if direction == "CALL" else "rs_negative"
        # direction-specific gate names
        htf_gate  = "htf_trend_bullish"  if direction == "CALL" else "htf_trend_bearish"
        macd_gate = "macd_bull_cross"    if direction == "CALL" else "macd_bear_cross"
        vwap_gate = "above_vwap"         if direction == "CALL" else "below_vwap"
        oi_gate   = "put_writing_or_short_covering" if direction == "CALL" else "call_writing_or_short_buildup"
        pcr_gate  = "pcr_supportive"     if direction == "CALL" else "pcr_bearish"

        dt = datetime.fromisoformat(sig["timestamp"])
        conf_val = sig.get("confidence") or sig.get("grade") or ""
        conf_str = conf_val if isinstance(conf_val, str) else f"{conf_val}"
        gate_score = sig.get("confidence") if isinstance(sig.get("confidence"), (int, float)) else None
        # In the DB, 'confidence' is the int gate score and 'grade' is A+/A-/B
        # In the CSV, 'confidence' is A+/A-/B and 'gate_score' is the numeric score
        grade     = sig.get("grade") or conf_str
        gate_score_num = sig.get("confidence") if isinstance(sig.get("confidence"), (int, float)) else ""

        def _cv(col_idx, value, bold=False, fg=_FG_WHITE, align="left"):
            c = ws.cell(row=r_idx, column=col_idx, value=value)
            c.font      = _font(bold=bold, colour=fg)
            c.fill      = _fill(bg)
            c.alignment = _align(align)
            c.border    = _border()

        col = 1
        _cv(col, dt.strftime("%d %b %Y")); col += 1
        _cv(col, dt.strftime("%H:%M")); col += 1
        _cv(col, sig.get("symbol", ""), bold=True, fg=_FG_WHITE); col += 1

        # Direction with colour
        dir_fg = _FG_GREEN if direction == "CALL" else _FG_RED
        c = ws.cell(row=r_idx, column=col, value=direction)
        c.font = _font(bold=True, colour=dir_fg); c.fill = _fill(bg)
        c.alignment = _align("center"); c.border = _border(); col += 1

        # Confidence grade
        grade_fg = {"A+": _FG_GREEN, "A-": _FG_AMBER, "B": _FG_MUTED}.get(grade, _FG_WHITE)
        c = ws.cell(row=r_idx, column=col, value=grade)
        c.font = _font(bold=True, colour=grade_fg); c.fill = _fill(bg)
        c.alignment = _align("center"); c.border = _border(); col += 1

        # Gate score
        _cv(col, gate_score_num, align="center"); col += 1

        # ── Gate columns ──────────────────────────────────────────────────
        gates_to_show = [
            "regime_supportive",
            rs_gate,
            "rsi_divergence",
            htf_gate,
            macd_gate,
            vwap_gate,
            "volume_expansion",
            oi_gate,
            pcr_gate,
        ]
        for gate_key in gates_to_show:
            passed = gate_key in active_gates if active_gates else None
            _gate_cell(ws, r_idx, col, passed)
            col += 1

        # ── Indicator values ──────────────────────────────────────────────
        _cv(col, sig.get("regime_type") or ""); col += 1
        _cv(col, sig.get("vix"), align="center"); col += 1
        _cv(col, round(float(sig["rsi"]), 1) if sig.get("rsi") else "", align="center"); col += 1
        _cv(col, round(float(sig["macd_hist"]), 4) if sig.get("macd_hist") else "", align="center"); col += 1
        _cv(col, round(float(sig["vwap"]), 2) if sig.get("vwap") else "", align="center"); col += 1
        _cv(col, round(float(sig["pcr"]), 2) if sig.get("pcr") else "", align="center"); col += 1
        _cv(col, sig.get("oi") or ""); col += 1
        htf_fg = _FG_GREEN if sig.get("htf_trend") == "BULLISH" else (_FG_RED if sig.get("htf_trend") == "BEARISH" else _FG_MUTED)
        c = ws.cell(row=r_idx, column=col, value=sig.get("htf_trend") or "")
        c.font = _font(colour=htf_fg); c.fill = _fill(bg); c.alignment = _align("center"); c.border = _border(); col += 1
        div_val = "Yes" if sig.get("divergence") else "No"
        div_fg  = _FG_GREEN if sig.get("divergence") else _FG_MUTED
        c = ws.cell(row=r_idx, column=col, value=div_val)
        c.font = _font(colour=div_fg); c.fill = _fill(bg); c.alignment = _align("center"); c.border = _border(); col += 1

        pattern = sig.get("candle_pattern") or ""
        pat_fg  = _FG_AMBER if pattern else _FG_MUTED
        c = ws.cell(row=r_idx, column=col, value=pattern if pattern else "—")
        c.font = _font(bold=bool(pattern), colour=pat_fg); c.fill = _fill(bg)
        c.alignment = _align("center"); c.border = _border(); col += 1

        # ── Entry/exit levels ─────────────────────────────────────────────
        _cv(col, sig.get("entry_spot"), align="center"); col += 1
        _cv(col, sig.get("sl_spot"), align="center"); col += 1
        _cv(col, sig.get("target1"), align="center"); col += 1
        _cv(col, sig.get("target2"), align="center"); col += 1
        _cv(col, sig.get("rr") or "", align="center"); col += 1
        _cv(col, sig.get("position_size") or ""); col += 1

        # ── Paper trade ───────────────────────────────────────────────────
        trades = trades_by_sid.get(sig["id"], [])
        trade  = trades[0] if trades else None

        if trade:
            outcome = trade.get("outcome") or ""
            pnl_fg  = _FG_GREEN if outcome == "WIN" else (_FG_RED if outcome == "LOSS" else _FG_AMBER)
            trade_bg = _BG_WIN if outcome == "WIN" else (_BG_LOSS if outcome == "LOSS" else bg)

            def _tv(c_idx, value, bold=False, fg=None, align="center"):
                fg = fg or _FG_WHITE
                c = ws.cell(row=r_idx, column=c_idx, value=value)
                c.font = _font(bold=bold, colour=fg); c.fill = _fill(trade_bg)
                c.alignment = _align(align); c.border = _border()

            _tv(col, trade.get("strike")); col += 1
            _tv(col, trade.get("expiry")); col += 1
            _tv(col, trade.get("option_type")); col += 1
            _tv(col, trade.get("lots")); col += 1
            _tv(col, trade.get("entry_premium")); col += 1
            # Invested amount = lots × entry_premium
            try:
                invested = round(float(trade["entry_premium"]) * float(trade["lots"]), 0) \
                    if trade.get("entry_premium") is not None and trade.get("lots") is not None else ""
            except (TypeError, ValueError):
                invested = ""
            c = ws.cell(row=r_idx, column=col, value=invested if invested != "" else "—")
            c.font = _font(bold=bool(invested), colour=_FG_AMBER if invested else _FG_MUTED)
            c.fill = _fill(trade_bg); c.alignment = _align("center"); c.border = _border(); col += 1
            _tv(col, trade.get("entry_spot")); col += 1
            entry_t = trade.get("entry_time", "")
            try:
                entry_t = datetime.fromisoformat(entry_t).strftime("%d %b %Y %H:%M")
            except Exception:
                pass
            _tv(col, entry_t); col += 1
            _tv(col, trade.get("exit_premium")); col += 1
            _tv(col, trade.get("exit_spot")); col += 1
            exit_t = trade.get("exit_time", "")
            try:
                exit_t = datetime.fromisoformat(exit_t).strftime("%d %b %Y %H:%M")
            except Exception:
                pass
            _tv(col, exit_t); col += 1
            exit_reason_fg = {"SL": _FG_RED, "T1": _FG_AMBER, "T2": _FG_GREEN, "TIME": _FG_MUTED}.get(trade.get("exit_reason",""), _FG_WHITE)
            c = ws.cell(row=r_idx, column=col, value=trade.get("exit_reason", ""))
            c.font = _font(bold=True, colour=exit_reason_fg); c.fill = _fill(trade_bg)
            c.alignment = _align("center"); c.border = _border(); col += 1
            _tv(col, round(float(trade["pnl_points"]), 2) if trade.get("pnl_points") is not None else ""); col += 1
            _tv(col, round(float(trade["pnl_rupees"]), 0) if trade.get("pnl_rupees") is not None else "", bold=True, fg=pnl_fg); col += 1
            pct = round(float(trade["pnl_percent"]), 2) if trade.get("pnl_percent") is not None else ""
            _tv(col, f"{pct}%" if pct != "" else "", fg=pnl_fg); col += 1
            c = ws.cell(row=r_idx, column=col, value=outcome)
            c.font = _font(bold=True, colour=pnl_fg); c.fill = _fill(trade_bg)
            c.alignment = _align("center"); c.border = _border(); col += 1
        else:
            for _ in range(16):
                c = ws.cell(row=r_idx, column=col, value="—")
                c.font = _font(colour=_FG_MUTED); c.fill = _fill(bg)
                c.alignment = _align("center"); c.border = _border()
                col += 1

        ws.row_dimensions[r_idx].height = 18


# ── Sheet 2: Gate Reference ───────────────────────────────────────────────────

_GATE_REF_COLS = [
    ("Gate Name",         18),
    ("Direction",          9),
    ("Weight (pts)",      12),
    ("Layer",              8),
    ("Indicator",         28),
    ("Timeframe",         11),
    ("Trigger Condition", 42),
    ("Why It Matters",    46),
]

_GATE_WHY = {
    "regime_supportive":              "Prevents trading into broad-market headwinds or event uncertainty. Gate of last resort — market must cooperate.",
    "rs_positive":                    "Picks stocks outperforming Nifty; ensures money is flowing INTO the stock. Avoids laggards.",
    "rs_negative":                    "Picks stocks underperforming Nifty; confirms structural selling pressure vs the index.",
    "rsi_divergence":                 "Highest-weight gate (30 pts). A divergence on 15-min signals genuine momentum shift, not noise. Strong edge in F&O.",
    "htf_trend_bullish":              "Daily trend acts as a tide. Trading with the daily EMA stack dramatically improves strike rate on intraday calls.",
    "htf_trend_bearish":              "Mirror of bullish. Selling into a downtrend is the high-probability intraday put setup.",
    "macd_bull_cross":                "Short-term (5-min) momentum confirmation. Cross means the fast average is now above slow — fresh buying momentum.",
    "macd_bear_cross":                "Fast average drops below slow on 5-min. Confirms distribution and downward momentum acceleration.",
    "above_vwap":                     "VWAP is the fair-value anchor. Price above it means buyers are in control intraday; call premiums compress less.",
    "below_vwap":                     "Price below VWAP signals intraday selling dominance. Put options benefit from accelerating downward momentum.",
    "volume_expansion":               "Confirms institutional participation. A signal with no volume is just retail noise. 1.5× average filters that out.",
    "put_writing_or_short_covering":  "Smart money shorting puts = they believe the downside is limited. Protective. Combined with call signals, increases conviction.",
    "call_writing_or_short_buildup":  "Institutions writing calls = they expect the upside to stall or reverse. Aligns with put bias. High-quality confirmation.",
    "pcr_supportive":                 "PCR > 1.2 means protective put buying dominates — market makers long delta → bullish lean for CALL setups.",
    "pcr_bearish":                    "PCR < 0.8 means call speculation dominates — skewed towards PUT setups; market is frothy and prone to mean-reverting down.",
}

_GATE_LAYER = {
    "regime_supportive": "1",
    "rs_positive": "2", "rs_negative": "2",
    "rsi_divergence": "3",
    "htf_trend_bullish": "3", "htf_trend_bearish": "3",
    "macd_bull_cross": "3", "macd_bear_cross": "3",
    "above_vwap": "3", "below_vwap": "3",
    "volume_expansion": "2/3",
    "put_writing_or_short_covering": "4", "call_writing_or_short_buildup": "4",
    "pcr_supportive": "4", "pcr_bearish": "4",
}


def _build_sheet2(wb):
    ws = wb.create_sheet("Gate Reference")
    ws.sheet_view.showGridLines = False

    _write_header_row(ws, _GATE_REF_COLS, row=1)

    for r_idx, (key, label, weight, direction, indicator, timeframe, threshold) in enumerate(_GATES, start=2):
        bg = _BG_EVEN if r_idx % 2 == 0 else _BG_ODD
        is_call = "CALL" in direction and "PUT" not in direction
        is_put  = "PUT"  in direction and "CALL" not in direction
        dir_fg  = _FG_GREEN if is_call else (_FG_RED if is_put else _FG_AMBER)

        data = [
            (label,                        _FG_WHITE, True,  "left"),
            (direction,                    dir_fg,    True,  "center"),
            (weight,                       _FG_AMBER, False, "center"),
            (_GATE_LAYER.get(key, ""),     _FG_PURPLE,False, "center"),
            (indicator,                    _FG_WHITE, False, "left"),
            (timeframe,                    _FG_MUTED, False, "center"),
            (threshold,                    _FG_WHITE, False, "left"),
            (_GATE_WHY.get(key, ""),       _FG_WHITE, False, "left"),
        ]
        for col_idx, (val, fg, bold, align) in enumerate(data, start=1):
            c = ws.cell(row=r_idx, column=col_idx, value=val)
            c.font      = _font(bold=bold, colour=fg)
            c.fill      = _fill(bg)
            c.alignment = _align(align, wrap=True)
            c.border    = _border()
        ws.row_dimensions[r_idx].height = 42


# ── Sheet 3: Paper Trades ─────────────────────────────────────────────────────

_PT_COLS = [
    ("Trade #",     9),
    ("Date",       11),
    ("Entry Time", 17),
    ("Symbol",     13),
    ("Direction",  10),
    ("Strike",      9),
    ("Expiry",     11),
    ("Opt Type",    9),
    ("Lots",            8),
    ("Entry Prem.",    12),
    ("Invested Amt.\n(lots × prem)", 14),
    ("Entry Spot",     11),
    ("Exit Prem.",     11),
    ("Exit Spot",      11),
    ("Exit Time",      17),
    ("Exit Reason",    12),
    ("P&L Pts",        10),
    ("P&L ₹",         11),
    ("P&L %",           9),
    ("Outcome",        11),
]


def _build_sheet3(wb, trades_by_sid):
    ws = wb.create_sheet("Paper Trades")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    _write_header_row(ws, _PT_COLS, row=1)

    all_trades = []
    for trade_list in trades_by_sid.values():
        all_trades.extend(trade_list)
    all_trades.sort(key=lambda t: t.get("entry_time") or "")

    for r_idx, t in enumerate(all_trades, start=2):
        outcome = t.get("outcome") or ""
        bg = _BG_WIN if outcome == "WIN" else (_BG_LOSS if outcome == "LOSS" else _BG_EVEN)
        pnl_fg = _FG_GREEN if outcome == "WIN" else (_FG_RED if outcome == "LOSS" else _FG_AMBER)

        def _tv(col, val, bold=False, fg=_FG_WHITE, align="center"):
            c = ws.cell(row=r_idx, column=col, value=val)
            c.font = _font(bold=bold, colour=fg); c.fill = _fill(bg)
            c.alignment = _align(align); c.border = _border()

        entry_t = t.get("entry_time", "")
        exit_t  = t.get("exit_time", "")
        try: entry_t = datetime.fromisoformat(entry_t).strftime("%d %b %Y %H:%M")
        except Exception: pass
        try: exit_t  = datetime.fromisoformat(exit_t).strftime("%d %b %Y %H:%M")
        except Exception: pass

        date_str = entry_t[:11] if entry_t else ""
        dir_fg = _FG_GREEN if t.get("direction") == "CALL" else _FG_RED
        exit_reason_fg = {"SL": _FG_RED, "T1": _FG_AMBER, "T2": _FG_GREEN, "TIME": _FG_MUTED}.get(t.get("exit_reason",""), _FG_WHITE)

        _tv(1, r_idx - 1)
        _tv(2, date_str)
        _tv(3, entry_t)
        _tv(4, t.get("symbol", ""), bold=True, fg=_FG_WHITE, align="left")
        c = ws.cell(row=r_idx, column=5, value=t.get("direction",""))
        c.font = _font(bold=True, colour=dir_fg); c.fill = _fill(bg); c.alignment = _align("center"); c.border = _border()
        _tv(6, t.get("strike"))
        _tv(7, t.get("expiry"))
        _tv(8, t.get("option_type"))
        _tv(9, t.get("lots"))
        _tv(10, t.get("entry_premium"))
        # Invested amount = lots × entry_premium
        try:
            invested = round(float(t["entry_premium"]) * float(t["lots"]), 0) \
                if t.get("entry_premium") is not None and t.get("lots") is not None else ""
        except (TypeError, ValueError):
            invested = ""
        c = ws.cell(row=r_idx, column=11, value=invested if invested != "" else "—")
        c.font = _font(bold=bool(invested), colour=_FG_AMBER if invested else _FG_MUTED)
        c.fill = _fill(bg); c.alignment = _align("center"); c.border = _border()
        _tv(12, t.get("entry_spot"))
        _tv(13, t.get("exit_premium"))
        _tv(14, t.get("exit_spot"))
        _tv(15, exit_t)
        c = ws.cell(row=r_idx, column=16, value=t.get("exit_reason",""))
        c.font = _font(bold=True, colour=exit_reason_fg); c.fill = _fill(bg); c.alignment = _align("center"); c.border = _border()
        _tv(17, round(float(t["pnl_points"]), 2) if t.get("pnl_points") is not None else "")
        pnl_r = round(float(t["pnl_rupees"]), 0) if t.get("pnl_rupees") is not None else ""
        c = ws.cell(row=r_idx, column=18, value=pnl_r)
        c.font = _font(bold=True, colour=pnl_fg); c.fill = _fill(bg); c.alignment = _align("center"); c.border = _border()
        pct = round(float(t["pnl_percent"]), 2) if t.get("pnl_percent") is not None else ""
        _tv(19, f"{pct}%" if pct != "" else "", fg=pnl_fg)
        c = ws.cell(row=r_idx, column=20, value=outcome)
        c.font = _font(bold=True, colour=pnl_fg); c.fill = _fill(bg); c.alignment = _align("center"); c.border = _border()
        ws.row_dimensions[r_idx].height = 18


# ── Sheet 4: P&L Summary ──────────────────────────────────────────────────────

def _build_sheet4(wb, signals, trades_by_sid):
    ws = wb.create_sheet("P&L Summary")
    ws.sheet_view.showGridLines = False

    all_trades = [t for tl in trades_by_sid.values() for t in tl]

    wins        = [t for t in all_trades if t.get("outcome") == "WIN"]
    losses      = [t for t in all_trades if t.get("outcome") == "LOSS"]
    breakevens  = [t for t in all_trades if t.get("outcome") == "BREAKEVEN"]
    pending     = [t for t in all_trades if not t.get("outcome")]
    closed      = wins + losses + breakevens

    total_pnl   = sum(float(t.get("pnl_rupees") or 0) for t in all_trades)
    win_rate    = len(wins) / len(closed) * 100 if closed else 0

    def _row(label, value, label_fg=_FG_WHITE, val_fg=_FG_WHITE, bold_val=False):
        nonlocal _row_idx
        c1 = ws.cell(row=_row_idx, column=1, value=label)
        c1.font = _font(colour=label_fg); c1.fill = _fill(_BG_ODD); c1.alignment = _align("left"); c1.border = _border()
        c2 = ws.cell(row=_row_idx, column=2, value=value)
        c2.font = _font(bold=bold_val, colour=val_fg); c2.fill = _fill(_BG_EVEN); c2.alignment = _align("center"); c2.border = _border()
        _row_idx += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16

    def _invested(t):
        try:
            return float(t["entry_premium"]) * float(t["lots"]) \
                if t.get("entry_premium") is not None and t.get("lots") is not None else 0.0
        except (TypeError, ValueError):
            return 0.0

    total_invested = sum(_invested(t) for t in all_trades)

    _row_idx = 1

    # ── Overall Performance ───────────────────────────────────────────────────
    for col in range(1, 5):
        c = ws.cell(row=_row_idx, column=col, value="Overall Performance" if col == 1 else "")
        c.font = _font(bold=True, size=12, colour=_FG_PURPLE)
        c.fill = _fill(_BG_HEADER); c.border = _border()
    _row_idx += 1

    _row("Total Signals (DB)",     len(signals))
    _row("Total Paper Trades",     len(all_trades))
    _row("Closed Trades",          len(closed))
    _row("Open / Pending",         len(pending))
    _row("Wins",                   len(wins),       val_fg=_FG_GREEN,  bold_val=True)
    _row("Losses",                 len(losses),     val_fg=_FG_RED,    bold_val=True)
    _row("Breakevens",             len(breakevens), val_fg=_FG_AMBER)
    _row("Win Rate",               f"{win_rate:.1f}%", val_fg=_FG_GREEN if win_rate >= 50 else _FG_RED, bold_val=True)

    total_fg = _FG_GREEN if total_pnl >= 0 else _FG_RED
    _row("Total Invested (₹)",     f"₹ {total_invested:,.0f}", val_fg=_FG_AMBER, bold_val=True)
    _row("Total P&L (₹)",          f"₹ {total_pnl:,.0f}", val_fg=total_fg, bold_val=True)
    if total_invested > 0:
        overall_ret = total_pnl / total_invested * 100
        ret_fg = _FG_GREEN if overall_ret >= 0 else _FG_RED
        _row("Overall Return %",   f"{overall_ret:+.2f}%", val_fg=ret_fg, bold_val=True)

    if wins:
        avg_win = sum(float(t.get("pnl_rupees") or 0) for t in wins) / len(wins)
        _row("Avg Win (₹)",        f"₹ {avg_win:,.0f}", val_fg=_FG_GREEN)
    if losses:
        avg_loss = sum(float(t.get("pnl_rupees") or 0) for t in losses) / len(losses)
        _row("Avg Loss (₹)",       f"₹ {avg_loss:,.0f}", val_fg=_FG_RED)

    # ── By Exit Reason ────────────────────────────────────────────────────────
    _row_idx += 1
    for col in range(1, 5):
        c = ws.cell(row=_row_idx, column=col, value="By Exit Reason" if col == 1 else "")
        c.font = _font(bold=True, size=12, colour=_FG_PURPLE)
        c.fill = _fill(_BG_HEADER); c.border = _border()
    _row_idx += 1

    for reason in ["T2", "T1", "SL", "TIME"]:
        reason_trades = [t for t in all_trades if t.get("exit_reason") == reason]
        reason_pnl    = sum(float(t.get("pnl_rupees") or 0) for t in reason_trades)
        reason_fg = {"T2": _FG_GREEN, "T1": _FG_AMBER, "SL": _FG_RED, "TIME": _FG_MUTED}.get(reason, _FG_WHITE)
        _row(f"{reason} exits", f"{len(reason_trades)} trades  |  ₹ {reason_pnl:,.0f}", val_fg=reason_fg)

    # ── By Date ───────────────────────────────────────────────────────────────
    _row_idx += 1
    for col, hdr in enumerate(["Date", "Trades  |  W / L", "Invested ₹", "P&L ₹"], start=1):
        c = ws.cell(row=_row_idx, column=col, value=hdr)
        c.font = _font(bold=True, size=11, colour=_FG_PURPLE)
        c.fill = _fill(_BG_HEADER); c.alignment = _align("center"); c.border = _border()
    _row_idx += 1

    by_date = {}
    for t in all_trades:
        d = (t.get("entry_time") or "")[:10]
        by_date.setdefault(d, []).append(t)

    for d in sorted(by_date):
        day_trades   = by_date[d]
        day_pnl      = sum(float(t.get("pnl_rupees") or 0) for t in day_trades)
        day_invested = sum(_invested(t) for t in day_trades)
        day_wins     = sum(1 for t in day_trades if t.get("outcome") == "WIN")
        day_losses   = sum(1 for t in day_trades if t.get("outcome") == "LOSS")
        day_pnl_fg   = _FG_GREEN if day_pnl >= 0 else _FG_RED
        bg = _BG_EVEN if _row_idx % 2 == 0 else _BG_ODD

        c1 = ws.cell(row=_row_idx, column=1, value=d)
        c1.font = _font(bold=True, colour=_FG_WHITE); c1.fill = _fill(bg)
        c1.alignment = _align("center"); c1.border = _border()

        summary = f"{len(day_trades)} trades  |  {day_wins}W / {day_losses}L"
        c2 = ws.cell(row=_row_idx, column=2, value=summary)
        c2.font = _font(colour=_FG_WHITE); c2.fill = _fill(bg)
        c2.alignment = _align("center"); c2.border = _border()

        c3 = ws.cell(row=_row_idx, column=3, value=f"₹ {day_invested:,.0f}" if day_invested else "—")
        c3.font = _font(bold=True, colour=_FG_AMBER if day_invested else _FG_MUTED)
        c3.fill = _fill(bg); c3.alignment = _align("center"); c3.border = _border()

        c4 = ws.cell(row=_row_idx, column=4, value=f"₹ {day_pnl:,.0f}")
        c4.font = _font(bold=True, colour=day_pnl_fg)
        c4.fill = _fill(bg); c4.alignment = _align("center"); c4.border = _border()

        _row_idx += 1

    # ── Date totals row ───────────────────────────────────────────────────────
    if by_date:
        c1 = ws.cell(row=_row_idx, column=1, value="TOTAL")
        c1.font = _font(bold=True, colour=_FG_PURPLE); c1.fill = _fill(_BG_HEADER)
        c1.alignment = _align("center"); c1.border = _border()

        c2 = ws.cell(row=_row_idx, column=2, value=f"{len(all_trades)} trades  |  {len(wins)}W / {len(losses)}L")
        c2.font = _font(bold=True, colour=_FG_WHITE); c2.fill = _fill(_BG_HEADER)
        c2.alignment = _align("center"); c2.border = _border()

        c3 = ws.cell(row=_row_idx, column=3, value=f"₹ {total_invested:,.0f}" if total_invested else "—")
        c3.font = _font(bold=True, colour=_FG_AMBER); c3.fill = _fill(_BG_HEADER)
        c3.alignment = _align("center"); c3.border = _border()

        total_fg2 = _FG_GREEN if total_pnl >= 0 else _FG_RED
        c4 = ws.cell(row=_row_idx, column=4, value=f"₹ {total_pnl:,.0f}")
        c4.font = _font(bold=True, colour=total_fg2); c4.fill = _fill(_BG_HEADER)
        c4.alignment = _align("center"); c4.border = _border()


# ── Public API ────────────────────────────────────────────────────────────────

def generate_xlsx() -> bytes:
    """Build the workbook and return raw bytes."""
    signals, trades_by_sid = _load_db_signals_and_trades()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    _build_sheet1(wb, signals, trades_by_sid)
    _build_sheet2(wb)
    _build_sheet3(wb, trades_by_sid)
    _build_sheet4(wb, signals, trades_by_sid)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
